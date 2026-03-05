"""
Minisoft Packing Match Logic
For kits missing Packing Method - Minisoft records, find kits with the same
components and suggest copying their Minisoft packing data.

Matching priority:
  1. Exact: same component items AND same quantities
  2. Partial: similarity score (Jaccard item overlap 60% + qty similarity 40%)

For partial matches, packing is inferred via linear regression across all
peer kits that share the same component types, treating the varying
quantity item as the independent variable.
"""

import json
import math
import pandas as pd
from collections import Counter, defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── File paths ───────────────────────────────────────────────────────────────
FILE_WITH_MINISOFT = r'C:\Users\info\.claude\projects\C--Users-info-Documents-Dev\3f12f581-6d4a-46d6-b111-1b5443647750\tool-results\mcp-claude_ai_NetSuite-ns_runCustomSuiteQL-1772732565024.txt'
FILE_ALL_MEMBERS   = r'C:\Users\info\.claude\projects\C--Users-info-Documents-Dev\3f12f581-6d4a-46d6-b111-1b5443647750\tool-results\mcp-claude_ai_NetSuite-ns_runCustomSuiteQL-1772731442868.txt'
FILE_NO_MINISOFT   = r'C:\Users\info\.claude\projects\C--Users-info-Documents-Dev\3f12f581-6d4a-46d6-b111-1b5443647750\tool-results\mcp-claude_ai_NetSuite-ns_runCustomSuiteQL-1772726771217.txt'
FILE_CROSS_REF     = r'C:\Users\info\.claude\projects\C--Users-info-Documents-Dev\3f12f581-6d4a-46d6-b111-1b5443647750\tool-results\mcp-claude_ai_NetSuite-ns_runCustomSuiteQL-1772731819042.txt'
OUTPUT_FILE        = r'C:\Users\info\Documents\Dev\Output\Minisoft_Match_Suggestions_chair8_sample_v2.xlsx'
EXCLUDED_MATCH_KIT_NAMES = {'SC AMA KIT_OIL NEW'.upper()}
CHAIR_BOX_CAPACITY_BY_FAMILY = {
    'CHAMONIX': 8,
    'STREAM': 8,
}

def load_json(path):
    with open(path, encoding='utf-8') as f:
        raw = f.read()
    outer = json.loads(raw)
    return json.loads(outer[0]['text'])['data']

def get_pkg_type(record):
    return 'Pallet' if record.get('num_boxes') is not None else 'Box'

def is_excluded_source_kit(kit_id):
    return kit_name_map.get(kit_id, '').strip().upper() in EXCLUDED_MATCH_KIT_NAMES

def get_kit_pack_type(kit_id):
    """
    Return the active packing style for a kit:
      - Box if any Box records exist (current style)
      - Pallet otherwise if pallet records exist
      - None if no records are present
    """
    recs = minisoft_by_kit.get(kit_id, [])
    if any(r.get('num_boxes') is None for r in recs):
        return 'Box'
    if any(r.get('num_boxes') is not None for r in recs):
        return 'Pallet'
    return None

# ── Load data ────────────────────────────────────────────────────────────────
print("Loading data...")
rows_with    = load_json(FILE_WITH_MINISOFT)
rows_all     = load_json(FILE_ALL_MEMBERS)
rows_missing = load_json(FILE_NO_MINISOFT)
rows_xref    = load_json(FILE_CROSS_REF)

xref_kit_ids = {r['item_id'] for r in rows_xref}
print(f"Kits with active cross references: {len(xref_kit_ids)}")

# ── Missing kits (cross-referenced, Kit/Package type only) ───────────────────
missing_kits = {
    r['id']: r['itemid']
    for r in rows_missing
    if r['itemtype'] in ('Kit', 'Package') and r['id'] in xref_kit_ids
}
missing_kit_ids = set(missing_kits.keys())
print(f"Kits missing Minisoft (with cross refs): {len(missing_kit_ids)}")

# ── All kit members ───────────────────────────────────────────────────────────
# kit_components_ids:   kit_id → frozenset of component item IDs (for set matching)
# kit_components_qty:   kit_id → frozenset of (item_id, qty) tuples (for qty matching)
# kit_components_label: kit_id → "name x qty, name x qty, ..." string
kit_components_ids   = defaultdict(set)
kit_components_qty   = defaultdict(set)
kit_name_map         = {}
comp_name_map        = {}

for row in rows_all:
    pid = row['parentitem']
    cid = row['component_item']
    qty = row['quantity']
    kit_components_ids[pid].add(cid)
    kit_components_qty[pid].add((cid, qty))
    kit_name_map[pid] = row['kit_name']
    comp_name_map[cid] = row['component_name']

kit_components_ids = {k: frozenset(v) for k, v in kit_components_ids.items()}
kit_components_qty = {k: frozenset(v) for k, v in kit_components_qty.items()}

def comp_label(kit_id):
    pairs = sorted(
        (comp_name_map.get(cid, str(cid)), qty)
        for cid, qty in kit_components_qty.get(kit_id, set())
    )
    return ', '.join(f'{n} x{q}' for n, q in pairs)

# ── Build Minisoft records for source kits ───────────────────────────────────
seen_minisoft            = {}
kit_minisoft_records     = defaultdict(list)
kit_with_minisoft_ids    = defaultdict(set)   # item IDs only
kit_with_minisoft_qty    = defaultdict(set)   # (item_id, qty) tuples

for row in rows_with:
    pid = row['parentitem']
    mid = row['minisoft_id']
    cid = row['component_item']
    qty = row['quantity']
    kit_with_minisoft_ids[pid].add(cid)
    kit_with_minisoft_qty[pid].add((cid, qty))
    kit_name_map[pid] = row['kit_name']
    comp_name_map[cid] = row['component_name']
    if mid not in seen_minisoft:
        seen_minisoft[mid] = {
            'minisoft_id' : mid,
            'created'     : row['created'],
            'pkg_number'  : row['custrecord4'],   # custrecord4 = Package Number shown in UI
            'num_boxes'   : row['custrecord5'],
            'weight'      : row['custrecord6'],
            'length'      : row['custrecord7'],
            'width'       : row['custrecord8'],
            'height'      : row['custrecord9'],
        }
        kit_minisoft_records[pid].append(seen_minisoft[mid])

kit_with_minisoft_ids = {k: frozenset(v) for k, v in kit_with_minisoft_ids.items()}
kit_with_minisoft_qty = {k: frozenset(v) for k, v in kit_with_minisoft_qty.items()}

minisoft_by_kit = {
    kid: sorted(recs, key=lambda r: (r['pkg_number'] or 0, r['minisoft_id']))
    for kid, recs in kit_minisoft_records.items()
}
kits_with_minisoft = set(minisoft_by_kit.keys())
print(f"Kits WITH Minisoft: {len(kits_with_minisoft)}")

# ── Exact lookup: qty_set → kit list ─────────────────────────────────────────

qty_set_to_kits = defaultdict(list)
for kid in kits_with_minisoft:
    if is_excluded_source_kit(kid):
        continue
    qset = kit_with_minisoft_qty.get(kid, frozenset())
    if qset:
        qty_set_to_kits[qset].append(kid)

# Pre-build dicts for fast scoring: kit_id → {item_id: qty}
def to_qty_dict(qty_set):
    return {item: qty for item, qty in qty_set}

source_qty_dicts = {
    kid: to_qty_dict(kit_with_minisoft_qty.get(kid, frozenset()))
    for kid in kits_with_minisoft
    if not is_excluded_source_kit(kid)
}

MIN_MATCH_SCORE = 0.40   # Below this → "No match" rather than a poor partial

def similarity_score(target_dict, source_dict):
    """
    Score 0-1 measuring how well a source kit matches a target.
      - Target recall    (45%): fraction of target component types covered by source.
                                Drops sharply when the source is missing target items.
      - Source precision (25%): fraction of source component types that exist in target.
                                Penalises sources that carry many unrelated extra components.
      - Quantity similarity (30%): closeness of quantities for shared items.
    Returns 0 if no components overlap.
    """
    shared = set(target_dict) & set(source_dict)
    if not shared:
        return 0.0
    target_recall    = len(shared) / len(target_dict)
    source_precision = len(shared) / len(source_dict)
    qty_sims = [min(target_dict[i], source_dict[i]) / max(target_dict[i], source_dict[i])
                for i in shared]
    avg_qty_sim = sum(qty_sims) / len(qty_sims)
    return target_recall * 0.45 + source_precision * 0.25 + avg_qty_sim * 0.30

# ── Peer index for inference: component_ids_frozenset → [source kit IDs] ─────
peer_index = defaultdict(list)
for kid in kits_with_minisoft:
    if is_excluded_source_kit(kid):
        continue
    ids_key = kit_with_minisoft_ids.get(kid, frozenset())
    if ids_key:
        peer_index[ids_key].append(kid)
print(f"Peer groups (component type sets): {len(peer_index)}")

def get_kit_totals(kit_id):
    """
    Return (total_weight, total_boxes) for a kit using the current packing method.

    Two packing styles exist in the data:
      - Box-style:   individual Box records (custrecord5 is None).
                     Each record = 1 box.  Weight = sum of Box weights.
      - Pallet-style: Pallet records (custrecord5 is not None).
                     Each record holds custrecord5 boxes.  Weight = sum of Pallet weights.

    When a kit has BOTH styles (old Pallet records + newer Box records), we use
    only the Box records because they reflect the current packing method.
    """
    recs        = minisoft_by_kit.get(kit_id, [])
    box_recs    = [r for r in recs if r['num_boxes'] is None]
    pallet_recs = [r for r in recs if r['num_boxes'] is not None]

    if box_recs:
        total_boxes  = len(box_recs)
        total_weight = sum(float(r['weight']) for r in box_recs if r['weight'] is not None)
    else:
        total_boxes  = sum(int(r['num_boxes']) for r in pallet_recs)
        total_weight = sum(float(r['weight']) for r in pallet_recs if r['weight'] is not None)

    return total_weight, total_boxes

def _pkg_sort_key(record):
    pkg = record.get('pkg_number')
    try:
        pkg_num = float(pkg) if pkg is not None else float('inf')
    except (TypeError, ValueError):
        pkg_num = float('inf')
    return (pkg_num, record.get('minisoft_id', 0))

def select_output_records(matched_id, packing_basis, inferred_boxes, match_type):
    """
    Pick source Minisoft rows to emit for output.
    - Use only the active packing basis (Box or Pallet).
    - For partial Box matches, limit emitted rows to inferred box count and
      prioritize package numbers 1..N so output appears as qty 1, qty 2, qty 3...
    """
    recs = list(minisoft_by_kit.get(matched_id, []))
    if not recs:
        return recs

    if packing_basis == 'Box':
        recs = [r for r in recs if r.get('num_boxes') is None]
    elif packing_basis == 'Pallet':
        pallet_recs = [r for r in recs if r.get('num_boxes') is not None]
        recs = pallet_recs if pallet_recs else recs

    recs = sorted(recs, key=_pkg_sort_key)

    if packing_basis == 'Box' and str(match_type).startswith('Partial'):
        try:
            target_boxes = int(inferred_boxes)
        except (TypeError, ValueError):
            target_boxes = 0
        if target_boxes > 0:
            by_pkg = {}
            for r in recs:
                pkg = r.get('pkg_number')
                if pkg is None:
                    continue
                try:
                    key = int(float(pkg))
                except (TypeError, ValueError):
                    continue
                if key not in by_pkg:
                    by_pkg[key] = r

            if by_pkg:
                chosen = [by_pkg[p] for p in sorted(by_pkg) if p <= target_boxes]
                if len(chosen) < target_boxes:
                    used_ids = {r.get('minisoft_id') for r in chosen}
                    for r in recs:
                        if r.get('minisoft_id') in used_ids:
                            continue
                        chosen.append(r)
                        used_ids.add(r.get('minisoft_id'))
                        if len(chosen) >= target_boxes:
                            break
                if chosen:
                    recs = chosen[:target_boxes]
            else:
                recs = recs[:target_boxes]

    return recs

def linear_predict(x_vals, y_vals, x_target):
    """
    Linear regression prediction for x_target given (x_vals, y_vals) data.
    Falls back to proportional scaling when only one data point is available.
    """
    n = len(x_vals)
    if n < 2:
        return y_vals[0] * x_target / x_vals[0] if x_vals[0] else y_vals[0]
    x_mean = sum(x_vals) / n
    y_mean = sum(y_vals) / n
    num = sum((x_vals[i] - x_mean) * (y_vals[i] - y_mean) for i in range(n))
    den = sum((x_vals[i] - x_mean) ** 2 for i in range(n))
    if den == 0:
        return y_mean
    slope = num / den
    return slope * x_target + (y_mean - slope * x_mean)

def get_effective_box_count_for_variable(matched_kit_id, target_dict, matched_dict):
    """
    For Box-style kits, estimate how many source boxes are tied to the shared
    variable component when target is only a subset of source components.

    Heuristic:
      - If target components are a strict subset of matched components,
        and box weights show a repeated modal weight, use that modal count
        as effective boxes for variable-component scaling.
      - Otherwise use all box records.
    """
    recs = minisoft_by_kit.get(matched_kit_id, [])
    box_recs = [r for r in recs if r.get('num_boxes') is None]
    total_boxes = len(box_recs)
    if total_boxes <= 1:
        return total_boxes, None

    target_items = set(target_dict)
    matched_items = set(matched_dict)
    is_strict_subset = target_items.issubset(matched_items) and len(target_items) < len(matched_items)
    if not is_strict_subset:
        return total_boxes, None

    weights = []
    for r in box_recs:
        w = r.get('weight')
        if w is None:
            return total_boxes, None
        try:
            weights.append(round(float(w), 1))
        except (TypeError, ValueError):
            return total_boxes, None

    counts = Counter(weights)
    modal_weight, modal_count = counts.most_common(1)[0]
    if modal_count >= 2 and modal_count < total_boxes:
        note = (
            f'Subset heuristic: using modal box weight {modal_weight:.1f}lbs '
            f'({modal_count}/{total_boxes} boxes)'
        )
        return modal_count, note

    return total_boxes, None

def infer_chair_family_boxes(target_dict):
    """
    Override inference for configured chair families.
    Counts family chair units and leg units, then uses max(chair, legs) so
    "including legs" is respected without double-counting.
    """
    total_boxes = 0
    notes = []

    all_leg_qty = 0
    for item_id, qty in target_dict.items():
        name = comp_name_map.get(item_id, '').upper()
        if 'LEG' in name:
            all_leg_qty += qty

    for family, cap in CHAIR_BOX_CAPACITY_BY_FAMILY.items():
        fam_chair_qty = 0

        for item_id, qty in target_dict.items():
            name = comp_name_map.get(item_id, '').upper()
            if family not in name:
                continue
            if any(tok in name for tok in ('CHAIR', 'SIDE', 'BUCKET', 'ARM', 'SOFA')):
                fam_chair_qty += qty

        if fam_chair_qty <= 0:
            continue

        # Legs are packed as their own box run; include them when family chairs exist.
        chair_boxes = math.ceil(fam_chair_qty / cap)
        leg_boxes = math.ceil(all_leg_qty / cap) if all_leg_qty > 0 else 0
        fam_boxes = chair_boxes + leg_boxes
        total_boxes += fam_boxes
        notes.append(
            f'{family}: chairs={fam_chair_qty} => {chair_boxes} boxes, '
            f'legs={all_leg_qty} => {leg_boxes} boxes, cap={cap}/box => {fam_boxes} boxes'
        )

    if total_boxes > 0:
        return total_boxes, 'Chair rule override: ' + '; '.join(notes)
    return None

def infer_packing(target_kit_id, matched_kit_id):
    """
    Infer total boxes and weight for a target kit by proportional scaling from the
    matched kit, using the shared component with the biggest quantity difference as
    the scaling axis.

    Example: target has 2 chairs, matched kit has 6 chairs packed in 3 boxes
    → 2 chairs per box → target needs 1 box.

    When multiple peer source kits with the same component types are available,
    linear regression across them is used instead for higher accuracy.
    """
    if not matched_kit_id:
        return None

    target_dict  = to_qty_dict(kit_components_qty.get(target_kit_id, frozenset()))
    matched_dict = source_qty_dicts.get(matched_kit_id, {})

    if not target_dict or not matched_dict:
        return None

    # User-provided chair-family box capacities override generic scaling logic.
    chair_override = infer_chair_family_boxes(target_dict)
    if chair_override:
        override_boxes, override_note = chair_override
        pred_weight = None
        matched_weight, matched_boxes = get_kit_totals(matched_kit_id)
        if matched_boxes > 0 and matched_weight > 0:
            pred_weight = matched_weight * (override_boxes / matched_boxes)
        return {
            'inferred_weight': round(pred_weight, 1) if pred_weight and pred_weight > 0 else None,
            'inferred_boxes' : override_boxes,
            'inference_notes': override_note,
        }

    # Variable item: shared component with the biggest absolute quantity difference.
    # This is the item whose packaging we're scaling.
    variable_item = None
    max_diff = 0
    for item in target_dict:
        if item in matched_dict:
            diff = abs(target_dict[item] - matched_dict[item])
            if diff > max_diff:
                max_diff = diff
                variable_item = item

    # If all shared items have the same qty, fall back to the highest-qty shared item
    if variable_item is None:
        shared = set(target_dict) & set(matched_dict)
        if shared:
            variable_item = max(shared, key=lambda x: target_dict[x])
        else:
            return None  # No shared components at all

    target_x  = target_dict[variable_item]
    matched_x = matched_dict.get(variable_item, 0)
    if matched_x == 0 or target_x == 0:
        return None

    var_name = comp_name_map.get(variable_item, str(variable_item))
    matched_pack_type = get_kit_pack_type(matched_kit_id)

    # For Box packing, trust matched-kit units/box directly.
    # This preserves the operational rule: infer cartons from box capacity.
    if matched_pack_type == 'Box':
        matched_weight, matched_boxes = get_kit_totals(matched_kit_id)
        if matched_boxes <= 0:
            return None
        eff_boxes, eff_note = get_effective_box_count_for_variable(
            matched_kit_id, target_dict, matched_dict
        )
        if eff_boxes > 0:
            matched_boxes = eff_boxes
        ratio = target_x / matched_x
        pred_boxes = matched_boxes * ratio
        pred_weight = matched_weight * ratio if matched_weight > 0 else 0
        per_unit = matched_x / matched_boxes
        notes = (
            f'Box capacity: "{var_name}" — matched has {matched_x:.0f} units in '
            f'{matched_boxes:.0f} boxes ({per_unit:.1f} units/box) → '
            f'{target_x} units = {pred_boxes:.1f} boxes (rounded up)'
        )
        if eff_note:
            notes = f'{notes}. {eff_note}.'
        inferred_boxes = max(1, math.ceil(max(0.0, pred_boxes)))
        return {
            'inferred_weight': round(pred_weight, 1) if pred_weight > 0 else None,
            'inferred_boxes' : inferred_boxes,
            'inference_notes': notes,
        }

    # ── Try peer regression when multiple peers with the same component types exist ──
    target_ids = kit_components_ids.get(target_kit_id, frozenset())
    points = []
    for peer_id in peer_index.get(target_ids, []):
        if matched_pack_type and get_kit_pack_type(peer_id) != matched_pack_type:
            continue
        pdict = source_qty_dicts.get(peer_id, {})
        if variable_item not in pdict:
            continue
        weight, boxes = get_kit_totals(peer_id)
        if weight > 0 or boxes > 0:
            points.append((pdict[variable_item], weight, boxes))

    if len(points) >= 2:
        points.sort(key=lambda p: p[0])
        x_vals = [p[0] for p in points]
        pred_weight = linear_predict(x_vals, [p[1] for p in points], target_x)
        pred_boxes  = linear_predict(x_vals, [p[2] for p in points], target_x)
        data_pts = '; '.join(f'qty={px}: {pw:.0f}lbs,{pb:.0f}boxes' for px, pw, pb in points)
        notes = f'Regression on "{var_name}" ({len(points)} peers): [{data_pts}] → qty={target_x}'
    else:
        # ── Proportional scaling from matched kit ────────────────────────────────
        matched_weight, matched_boxes = get_kit_totals(matched_kit_id)
        if matched_weight == 0 and matched_boxes == 0:
            return None

        ratio = target_x / matched_x
        pred_boxes  = matched_boxes * ratio
        pred_weight = matched_weight * ratio

        per_unit = matched_x / matched_boxes if matched_boxes > 0 else 0
        notes = (
            f'Proportional: "{var_name}" — matched has {matched_x:.0f} units in '
            f'{matched_boxes:.0f} boxes ({per_unit:.1f} units/box) → '
            f'{target_x} units = {pred_boxes:.1f} boxes'
        )

    inferred_boxes = max(0, round(pred_boxes))

    return {
        'inferred_weight': round(pred_weight, 1) if pred_weight > 0 else None,
        'inferred_boxes' : inferred_boxes,
        'inference_notes': notes,
    }

# ── Match missing kits ────────────────────────────────────────────────────────
results = []

for kit_id in missing_kit_ids:
    kit_name   = missing_kits[kit_id]
    my_ids     = kit_components_ids.get(kit_id, frozenset())
    my_qty     = kit_components_qty.get(kit_id, frozenset())
    my_label   = comp_label(kit_id)

    if not my_ids:
        results.append({
            'Kit ID': kit_id, 'Kit Name': kit_name,
            'My Components': '', 'Match Type': 'No components found',
            'Matched Kit ID': '', 'Matched Kit Name': '', 'Source Components': '',
            'Match Score': '', 'Date Created': '', 'Package Number': '', 'Package Type': '',
            'Num Boxes in Pallet': '', 'Weight': '', 'Length': '', 'Width': '', 'Height': '',
            'Inferred Boxes': '', 'Inferred Weight': '', 'Inference Notes': '',
        })
        continue

    my_dict = to_qty_dict(my_qty)

    # 1. Exact: identical items + quantities
    exact = qty_set_to_kits.get(my_qty, [])
    if exact:
        match_type  = 'Exact'
        matched_id  = exact[0]
        best_score  = 1.0
    else:
        # 2. Score every source kit and pick the best
        best_score  = 0.0
        matched_id  = None
        for kid, src_dict in source_qty_dicts.items():
            score = similarity_score(my_dict, src_dict)
            if score > best_score:
                best_score = score
                matched_id = kid

        if matched_id and best_score >= MIN_MATCH_SCORE:
            match_type = 'Partial (similar components)'
        else:
            match_type = 'No match'
            matched_id = None

    # Infer packing for partial matches only
    inference = None
    if match_type.startswith('Partial'):
        inference = infer_packing(kit_id, matched_id)

    inferred_boxes  = inference['inferred_boxes']  if inference else ''
    inferred_weight = inference['inferred_weight'] if inference else ''
    inference_notes = inference['inference_notes'] if inference else ''
    packing_basis   = get_kit_pack_type(matched_id) if matched_id else ''

    if matched_id:
        matched_name   = kit_name_map.get(matched_id, str(matched_id))
        source_label   = comp_label(matched_id)
        score_pct = f'{best_score:.0%}'
        output_records = select_output_records(
            matched_id, packing_basis, inferred_boxes, match_type
        )
        for ms in output_records:
            results.append({
                'Kit ID'             : kit_id,
                'Kit Name'           : kit_name,
                'My Components'      : my_label,
                'Match Type'         : match_type,
                'Match Score'        : score_pct,
                'Matched Kit ID'     : matched_id,
                'Matched Kit Name'   : matched_name,
                'Source Components'  : source_label,
                'Date Created'       : ms['created'],
                'Package Number'     : ms['pkg_number'],
                'Package Type'       : get_pkg_type(ms),
                'Num Boxes in Pallet': ms['num_boxes'],
                'Weight'             : ms['weight'],
                'Length'             : ms['length'],
                'Width'              : ms['width'],
                'Height'             : ms['height'],
                'Inferred Boxes'     : inferred_boxes,
                'Inferred Weight'    : inferred_weight,
                'Packing Basis'      : packing_basis,
                'Inference Notes'    : inference_notes,
            })
    else:
        results.append({
            'Kit ID': kit_id, 'Kit Name': kit_name,
            'My Components': my_label, 'Match Type': match_type,
            'Match Score': '', 'Matched Kit ID': '', 'Matched Kit Name': '', 'Source Components': '',
            'Date Created': '', 'Package Number': '', 'Package Type': '',
            'Num Boxes in Pallet': '', 'Weight': '', 'Length': '', 'Width': '', 'Height': '',
            'Inferred Boxes': '', 'Inferred Weight': '', 'Packing Basis': '', 'Inference Notes': '',
        })

df = pd.DataFrame(results)

# ── Summary ───────────────────────────────────────────────────────────────────
match_counts = df.drop_duplicates('Kit ID')['Match Type'].value_counts()
print("\nMatch summary:")
print(match_counts.to_string())

# ── Write Excel ───────────────────────────────────────────────────────────────
with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    summary_df = match_counts.reset_index()
    summary_df.columns = ['Match Type', 'Kit Count']
    summary_df.to_excel(writer, sheet_name='Summary', index=False)
    df[df['Match Type'] == 'Exact'].sort_values('Kit Name').to_excel(
        writer, sheet_name='Exact Matches', index=False)
    df[df['Match Type'].str.startswith('Partial', na=False)].sort_values('Kit Name').to_excel(
        writer, sheet_name='Partial Matches', index=False)
    df[df['Match Type'] == 'No match'].sort_values('Kit Name').to_excel(
        writer, sheet_name='No Match', index=False)
    df.sort_values(['Match Type', 'Kit Name']).to_excel(writer, sheet_name='All', index=False)

# ── Style ─────────────────────────────────────────────────────────────────────
wb = load_workbook(OUTPUT_FILE)
COLORS = {'Exact': 'C6EFCE', 'Partial': 'FFEB9C', 'No match': 'FFC7CE', 'header': '1F4E79'}
header_font = Font(color='FFFFFF', bold=True)

for ws in wb.worksheets:
    for cell in ws[1]:
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    if 'Match Type' in [c.value for c in ws[1]]:
        mt_col = next(c.column for c in ws[1] if c.value == 'Match Type')
        for row in ws.iter_rows(min_row=2):
            mt_val = row[mt_col - 1].value or ''
            color = COLORS['Exact'] if 'Exact' in mt_val else \
                    COLORS['Partial'] if 'Partial' in mt_val else \
                    COLORS['No match'] if 'No match' in mt_val else 'F2F2F2'
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            for cell in row:
                cell.fill = fill
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

wb.save(OUTPUT_FILE)
print(f'\nDone! Output: {OUTPUT_FILE}')
