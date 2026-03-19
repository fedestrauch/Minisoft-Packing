"""
delivery_comparison.py
----------------------
Parse ABS Transport invoices (PDFs) from a folder, look up the matching
Sales Order in NetSuite, and compare:
  - Amount Quoted  = total on the PDF invoice (what ABS charged IHM)
  - Amount Charged = Delivery Charges line on the NS Sales Order (what IHM charged the customer)

Output: delivery_comparison.xlsx

Usage:
    python delivery_comparison.py [folder]

Default folder: C:\\Users\\info\\Documents\\Dev\\Quote-Invoice
"""

import os
import re
import sys
import warnings
from datetime import datetime

import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings("ignore")

# ── Config ───────────────────────────────────────────────────────────────────
FOLDER  = r"C:\Users\info\Documents\Dev\Quote-Invoice"
OUTPUT  = os.path.join(FOLDER, "delivery_comparison.xlsx")

# NetSuite item internal ID for "Delivery Charges"
NS_DELIVERY_ITEM_ID = "7751"


# ── PDF parsing ──────────────────────────────────────────────────────────────

def parse_pdf(path: str) -> dict:
    """
    Extract from one ABS Transport PDF:
      - inv_no      : invoice number (from filename or "Invoice No:" field)
      - date        : delivery/service date from the notes line  (e.g. "3/2/26")
      - amount      : Balance Due / Total amount
      - so_number   : SO number referenced in the notes  (e.g. "SO203340")
    """
    with pdfplumber.open(path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    fname = os.path.basename(path)

    # Invoice number — from filename or header
    inv_m = re.search(r"Invoice No:\s*(\S+)", text, re.I)
    inv_no = inv_m.group(1) if inv_m else os.path.splitext(fname)[0]

    # Amount — prefer Balance Due, fall back to last Total value
    bal_m = re.search(r"Balance Due\s+\$?([\d,]+\.?\d*)", text, re.I)
    tot_m = re.findall(r"Total\s+\$?([\d,]+\.?\d*)", text, re.I)
    if bal_m:
        amount = float(bal_m.group(1).replace(",", ""))
    elif tot_m:
        amount = float(tot_m[-1].replace(",", ""))
    else:
        amount = None

    # Delivery / service date — first date in the notes line after Balance Due
    # Pattern: "3/2/26 - delivery to ..."  or  "3/5/26 - DELIVERY ..."
    notes_m = re.search(
        r"Balance Due.*?\n(\d{1,2}/\d{1,2}/\d{2,4})\s+-",
        text, re.I | re.DOTALL
    )
    if notes_m:
        raw_date = notes_m.group(1)
        # normalise 2-digit year → 4-digit
        try:
            dt = datetime.strptime(raw_date, "%m/%d/%y")
        except ValueError:
            dt = datetime.strptime(raw_date, "%m/%d/%Y")
        date_str = dt.strftime("%-m/%-d/%Y") if os.name != "nt" else dt.strftime("%#m/%#d/%Y")
    else:
        # fall back to header Date field
        hdr_m = re.search(r"\bDate:\s*(\d+/\d+/\d+)", text)
        date_str = hdr_m.group(1) if hdr_m else None

    # SO number — "INVOICE # SO203340" or "INVOICE # S0203326" (zero variant)
    so_m = re.search(r"INVOICE\s*#\s*(S[O0]\d+)", text, re.I)
    if so_m:
        raw_so = so_m.group(1)
        # normalise S0XXXXXX → SOXXXXXX
        so_number = re.sub(r"^S0", "SO", raw_so, flags=re.I).upper()
    else:
        so_number = None

    return {
        "inv_no":    inv_no,
        "date":      date_str,
        "amount":    amount,
        "so_number": so_number,
        "file":      fname,
    }


# ── NetSuite lookup ──────────────────────────────────────────────────────────

def lookup_ns(so_numbers: list[str]) -> dict[str, dict]:
    """
    Batch query NetSuite for a list of SO tranids.
    Returns dict keyed by tranid (uppercase) →
        {"customer": str, "delivery_charge": float|None, "ns_tranid": str}
    """
    try:
        from netsuite_client import query as ns_query
    except ImportError:
        print("  WARNING: netsuite_client not available — NS columns will be empty.")
        return {}

    if not so_numbers:
        return {}

    # Build quoted list for IN clause
    quoted = ", ".join(f"'{s}'" for s in so_numbers)

    # 1. Get SO IDs and customer names
    try:
        so_rows = ns_query(f"""
            SELECT t.id, t.tranid, BUILTIN.DF(t.entity) AS customer_name
            FROM transaction t
            WHERE t.tranid IN ({quoted})
              AND t.type = 'SalesOrd'
        """)
    except Exception as e:
        print(f"  NS SO lookup error: {e}")
        return {}

    if not so_rows:
        print("  WARNING: no SOs found in NetSuite for the given SO numbers.")
        return {}

    # Map internal_id → tranid and customer
    id_to_tranid   = {r["id"]: r["tranid"].upper() for r in so_rows}
    tranid_to_cust = {r["tranid"].upper(): r["customer_name"] for r in so_rows}
    so_internal_ids = ", ".join(id_to_tranid.keys())

    # 2. Get Delivery Charges line for each SO
    try:
        line_rows = ns_query(f"""
            SELECT tl.transaction, tl.rate
            FROM transactionLine tl
            WHERE tl.transaction IN ({so_internal_ids})
              AND tl.item = {NS_DELIVERY_ITEM_ID}
        """)
    except Exception as e:
        print(f"  NS line lookup error: {e}")
        line_rows = []

    # Sum delivery charge per SO (in case there are multiple lines)
    delivery_by_id: dict[str, float] = {}
    for row in line_rows:
        tid = row["transaction"]
        val = float(row["rate"] or 0)
        delivery_by_id[tid] = delivery_by_id.get(tid, 0.0) + val

    # Build final result
    result = {}
    for ns_id, tranid in id_to_tranid.items():
        result[tranid] = {
            "customer":        tranid_to_cust.get(tranid, ""),
            "delivery_charge": delivery_by_id.get(ns_id),
            "ns_tranid":       tranid,
        }

    return result


# ── Output ───────────────────────────────────────────────────────────────────

def write_excel(rows: list[dict], path: str):
    cols = ["SO #", "Date", "Customer", "Amount Quoted ($)", "Amount Charged ($)", "Difference ($)", "Invoice #"]
    df = pd.DataFrame(rows, columns=cols)
    df.to_excel(path, index=False, engine="openpyxl")

    wb = load_workbook(path)
    ws = wb.active

    # Styles
    HEADER_FILL = PatternFill("solid", fgColor="2E4057")  # dark navy
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
    GRN_FILL    = PatternFill("solid", fgColor="C6EFCE")
    BORDER      = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin",  color="CCCCCC"),
    )

    # Header row
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20

    # Data rows
    diff_col_idx = cols.index("Difference ($)")
    diff_col     = diff_col_idx + 1
    for i, row in enumerate(df.iterrows(), start=2):
        _, row = row
        diff = row.iloc[diff_col_idx]
        if diff is None:
            fill = PatternFill()
        elif diff > 0:
            fill = GRN_FILL  # IHM made more than it paid
        elif diff < 0:
            fill = RED_FILL  # IHM paid more than it charged
        else:
            fill = PatternFill()

        for cell in ws[i]:
            cell.border = BORDER
            if cell.column == diff_col and diff is not None:
                cell.fill = fill

    # Column widths
    widths = {
        "A": 14,  # SO #
        "B": 12,  # Date
        "C": 28,  # Customer
        "D": 18,  # Amount Quoted
        "E": 20,  # Amount Charged
        "F": 16,  # Difference
        "G": 16,  # Invoice #
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # Currency format for money columns
    for col_letter in ("D", "E", "F"):
        col_idx = ord(col_letter) - ord("A") + 1
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(path)


# ── Main ─────────────────────────────────────────────────────────────────────

def run(folder: str = FOLDER):
    pdf_files = sorted(
        f for f in os.listdir(folder)
        if f.lower().endswith(".pdf")
    )
    if not pdf_files:
        print(f"No PDFs found in {folder}")
        return

    print(f"Found {len(pdf_files)} PDF(s) in {folder}\n")

    # Parse all PDFs
    parsed = []
    for fname in pdf_files:
        path = os.path.join(folder, fname)
        data = parse_pdf(path)
        parsed.append(data)
        status = f"SO={data['so_number'] or '?':12s}  amount=${data['amount'] or '?'}"
        print(f"  {fname:20s}  {status}")

    # Collect unique SO numbers for batch NS lookup
    so_numbers = list({
        p["so_number"] for p in parsed
        if p["so_number"]
    })
    print(f"\nLooking up {len(so_numbers)} SO(s) in NetSuite...")
    ns_data = lookup_ns(so_numbers)

    # Build output rows
    rows = []
    for p in parsed:
        so = p["so_number"]
        ns = ns_data.get(so, {}) if so else {}

        quoted   = p["amount"]
        charged  = ns.get("delivery_charge")
        diff     = round(charged - quoted, 2) if (charged is not None and quoted is not None) else None

        rows.append({
            "SO #":                so or "NOT FOUND",
            "Date":                p["date"],
            "Customer":            ns.get("customer", ""),
            "Amount Quoted ($)":   quoted,
            "Amount Charged ($)":  charged,
            "Difference ($)":      diff,
            "Invoice #":           p["inv_no"],
        })

        flag = ""
        if diff is not None:
            flag = f"  diff=${diff:+.2f}"
        print(f"  {so or '?':12s}  customer={ns.get('customer','?'):<25s}  "
              f"quoted=${quoted or '?'}  charged=${charged or '?'}{flag}")

    # Write Excel
    write_excel(rows, OUTPUT)
    print(f"\nOutput: {OUTPUT}")


if __name__ == "__main__":
    folder = sys.argv[1] if len(sys.argv) > 1 else FOLDER
    run(folder)
