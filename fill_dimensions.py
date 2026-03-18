"""
fill_dimensions.py
------------------
Match 77 IHM items (Items.xlsx) against supplier dimension spreadsheets
using two-phase fuzzy matching, then write Items_filled.xlsx with
L_cm, W_cm, H_cm, GW_lbs and color-coded confidence.

Usage:
    python fill_dimensions.py

Requirements:
    pip install pandas openpyxl rapidfuzz
"""

import os
import re
import warnings
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from rapidfuzz import fuzz, process

warnings.filterwarnings("ignore", category=UserWarning)

FOLDER = r"C:\Users\info\Documents\Dev\Re_ Listado de Items sin Dimensiones de Cajas o Pallets"
TARGET_FILE = os.path.join(FOLDER, "Items.xlsx")
OUTPUT_FILE = os.path.join(FOLDER, "Items_filled.xlsx")

KG_TO_LBS    = 2.20462
CM_TO_IN     = 0.393701
SCORE_GREEN  = 75   # auto-fill (high confidence)
SCORE_YELLOW = 50   # flag for review

# Known IHM brand prefixes to strip from SKUs before matching
_BRAND_PREFIXES = {"SC", "IN", "PLI", "SCI", "IHM"}
_BOX_SUFFIX_RE  = re.compile(r"_B\d+$", re.I)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _to_float(val):
    """Convert value to float; return None for blank, zero, or unparseable."""
    if val is None:
        return None
    if isinstance(val, float) and (val != val):   # NaN check
        return None
    try:
        f = float(val)
        return f if f > 0 else None
    except (ValueError, TypeError):
        return None


def parse_dims(val):
    """
    Parse a dimension value to (L, W, H) in cm.

    Handles:
      - Tuple (already split by caller)                → passthrough
      - "63.5 x 91.5 x 16.0" or "63.5×91.5×16.0"     → (63.5, 91.5, 16.0)
      - "108.00x56.70x225.00"  (no spaces around x)   → (108.0, 56.7, 225.0)
      - "Panama table: 79.10\\nPanama Corner: 207.10"  → (79.1, 207.1, None)
      - None / NaN                                     → (None, None, None)
    """
    if isinstance(val, tuple):
        return val
    if val is None or (isinstance(val, float) and val != val):
        return (None, None, None)
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return (None, None, None)

    def _all_zero(lst):
        return bool(lst) and all(f == 0.0 for f in lst if f is not None)

    # Strategy 1: split on x / × separators (handles "LxWxH" and "L x W x H")
    parts = re.split(r"\s*[×xX]\s*", s)
    if len(parts) >= 2:
        nums = []
        for part in parts:
            m = re.findall(r"\d+\.?\d*", part)
            if m:
                nums.append(float(m[-1]))  # last number in the segment
        if len(nums) >= 2:
            while len(nums) < 3:
                nums.append(None)
            result = tuple(nums[:3])
            if not _all_zero([v for v in result if v is not None]):
                return result

    # Strategy 2: pick decimal numbers (for multipart label strings like
    # "Panama table: 79.10\nPanama Corner: 207.10")
    dec_nums = re.findall(r"\d+\.\d+", s)
    if dec_nums:
        floats = [float(n) for n in dec_nums[:3]]
        while len(floats) < 3:
            floats.append(None)
        if not _all_zero([v for v in floats if v is not None]):
            return tuple(floats)

    # Strategy 3: any integers as last resort
    int_nums = re.findall(r"\d+", s)
    floats = [float(n) for n in int_nums[:3]]
    while len(floats) < 3:
        floats.append(None)
    if _all_zero([v for v in floats if v is not None]):
        return (None, None, None)
    return tuple(floats)


def _find_col(df, patterns):
    """Return first column name that matches any regex pattern (case-insensitive)."""
    for pat in patterns:
        for col in df.columns:
            if re.search(pat, str(col), re.I):
                return col
    return None


def _clean_match_name(name):
    """Strip specification noise from source item names to keep only the product descriptor.

    'Palau corner chair, w/csh DCZ w/back csh 8824...' → 'palau corner chair'
    'Gala 3-seater sofa, w/csh OLV...'                → 'gala 3-seater sofa'
    """
    clean = re.split(r",\s*w/", str(name), maxsplit=1, flags=re.I)[0].strip()
    return clean.lower()


def _row_to_record(item_name, L, W, H, gw_kg=None, gw_lbs=None, source=""):
    """Build a normalised source record dict."""
    if gw_lbs is None and gw_kg is not None:
        gw_lbs = round(gw_kg * KG_TO_LBS, 2)
    return {
        "item_name": item_name,
        "L_cm":      L,
        "W_cm":      W,
        "H_cm":      H,
        "GW_lbs":    gw_lbs,
        "source":    source,
    }


# ── Source loaders ────────────────────────────────────────────────────────────

def load_teak():
    path = os.path.join(FOLDER, "Lista items 1 TEAK.xlsx")
    df = pd.read_excel(path, sheet_name=0)
    df.columns = [str(c).strip() for c in df.columns]

    col_name = _find_col(df, [r"item.?name", r"art.+name"])
    col_L    = _find_col(df, [r"carton.?l\s*\(cm\)", r"carton\s*l$"])
    col_W    = _find_col(df, [r"carton.?w\s*\(cm\)", r"carton\s*w$"])
    col_H    = _find_col(df, [r"carton.?h\s*\(cm\)", r"carton\s*h$"])
    col_lbs  = _find_col(df, [r"gw\s*\(lbs?\)"])
    col_kg   = _find_col(df, [r"gross.?weight\s*\(kg\)", r"gw\s*\(kg\)"])

    records = []
    for _, r in df.iterrows():
        name = str(r.get(col_name, "") or "").strip()
        if not name or name.lower() == "nan":
            continue
        L      = _to_float(r.get(col_L))
        W      = _to_float(r.get(col_W))
        H      = _to_float(r.get(col_H))
        gw_lbs = _to_float(r.get(col_lbs)) if col_lbs else None
        gw_kg  = _to_float(r.get(col_kg))  if col_kg  else None
        records.append(_row_to_record(name, L, W, H, gw_kg=gw_kg, gw_lbs=gw_lbs, source="TEAK"))
    return pd.DataFrame(records)


def load_new():
    path = os.path.join(FOLDER, "Lista items new.xlsx")
    df = pd.read_excel(path, sheet_name=0)
    df.columns = [str(c).strip() for c in df.columns]

    col_name = _find_col(df, [r"item.?name", r"art.+name"])
    col_L    = _find_col(df, [r"carton.?l\s*\(cm\)", r"carton\s*l$"])
    col_W    = _find_col(df, [r"carton.?w\s*\(cm\)", r"carton\s*w$"])
    col_H    = _find_col(df, [r"carton.?h\s*\(cm\)", r"carton\s*h$"])
    col_lbs  = _find_col(df, [r"gw\s*\(lbs?\)"])
    col_kg   = _find_col(df, [r"gross.?weight\s*\(kg\)", r"gw\s*\(kg\)"])

    records = []
    for _, r in df.iterrows():
        name = str(r.get(col_name, "") or "").strip()
        if not name or name.lower() == "nan":
            continue
        L      = _to_float(r.get(col_L))
        W      = _to_float(r.get(col_W))
        H      = _to_float(r.get(col_H))
        gw_lbs = _to_float(r.get(col_lbs)) if col_lbs else None
        gw_kg  = _to_float(r.get(col_kg))  if col_kg  else None
        records.append(_row_to_record(name, L, W, H, gw_kg=gw_kg, gw_lbs=gw_lbs, source="new"))
    return pd.DataFrame(records)


def _detect_header_row(raw_df, min_str_cols=8, max_scan=10):
    """Return index of the first row that looks like a header (many string values)."""
    for i in range(min(max_scan, len(raw_df))):
        row = raw_df.iloc[i]
        str_count = sum(1 for v in row if isinstance(v, str) and v.strip())
        if str_count >= min_str_cols:
            return i
    return 0


def load_new2():
    path = os.path.join(FOLDER, "Lista items new2.xlsx")
    all_sheets = pd.read_excel(path, sheet_name=None, header=None)

    records = []
    for sheet_name, raw in all_sheets.items():
        h = _detect_header_row(raw)
        df = raw.iloc[h + 1:].copy()
        # Normalise column names: strip whitespace and replace embedded newlines
        df.columns = [str(c).strip().replace("\n", " ") for c in raw.iloc[h].tolist()]
        df = df.reset_index(drop=True)

        col_name = _find_col(df, [r"art.+name", r"item.?name"])
        col_lbs  = _find_col(df, [r"gw\s*\(lbs?\)"])
        col_kg   = _find_col(df, [r"gross.?weight", r"gw\s*\(kg\)", r"^gw$"])

        # Try separate L/W/H columns first (TERRAZZA layout)
        col_L   = _find_col(df, [r"carton.?l\s*[\(/]", r"cartonl", r"^carton\s*l$"])
        col_W   = _find_col(df, [r"carton.?w\s*[\(/]", r"cartonw", r"^carton\s*w$"])
        col_H   = _find_col(df, [r"carton.?h\s*[\(/]", r"cartonh", r"^carton\s*h$"])
        # Fall back to combined string (COL/LSG layout)
        col_dim = _find_col(df, [r"carton.+l.?w.?h", r"carton.+dim", r"carton.+cm"])
        use_sep = bool(col_L and col_W and col_H)

        short = sheet_name[:12]
        for _, r in df.iterrows():
            name = str(r.get(col_name, "") or "").strip() if col_name else ""
            if not name or name.lower() == "nan":
                continue
            if use_sep:
                L = _to_float(r.get(col_L))
                W = _to_float(r.get(col_W))
                H = _to_float(r.get(col_H))
            elif col_dim:
                L, W, H = parse_dims(r.get(col_dim))
            else:
                L = W = H = None
            gw_lbs = _to_float(r.get(col_lbs)) if col_lbs else None
            gw_kg  = _to_float(r.get(col_kg))  if col_kg  else None
            records.append(_row_to_record(name, L, W, H, gw_kg=gw_kg, gw_lbs=gw_lbs, source=f"new2/{short}"))

    return pd.DataFrame(records)


def load_all_sources():
    dfs = []
    for loader, label in [(load_teak, "TEAK"), (load_new, "new"), (load_new2, "new2")]:
        try:
            df = loader()
            df = df[df["item_name"].notna() & (df["item_name"] != "")]
            df = df[df[["L_cm", "W_cm", "H_cm", "GW_lbs"]].notna().any(axis=1)]
            dfs.append(df)
            print(f"  Loaded {label}: {len(df)} usable rows")
        except Exception as e:
            print(f"  WARNING: failed to load {label}: {e}")
    result = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
    if not result.empty:
        result["_short_norm"] = result["item_name"].apply(_clean_match_name)
    return result


# ── SKU normalisation ─────────────────────────────────────────────────────────

def strip_box_suffix(sku):
    """Remove trailing B1/B2/B3 box-number suffix from an IHM SKU."""
    return _BOX_SUFFIX_RE.sub("", sku)


def normalize_sku(sku):
    """
    Convert an IHM SKU to a normalised string for fuzzy matching.

    "SC GALA_3SEATER_8839_CSH"  →  "gala 3seater csh"
    "IN_IC_974_LO_B2"           →  "ic 974 lo"
    "PLI_CURAZAO_CHAIR_WHT"     →  "curazao chair wht"
    """
    sku = strip_box_suffix(sku)
    tokens = re.split(r"[_\s]+", sku.strip())
    result = []
    for i, tok in enumerate(tokens):
        if not tok:
            continue
        # Strip known brand prefix (first token)
        if i == 0 and tok.upper() in _BRAND_PREFIXES:
            continue
        # Strip 4+-digit numeric style/color codes
        if re.fullmatch(r"\d{4,}", tok):
            continue
        result.append(tok.lower())
    return " ".join(result)


def _family_token(norm_sku):
    """Extract the first meaningful word (≥3 chars, not a number) from a normalised SKU."""
    for tok in norm_sku.split():
        if len(tok) >= 3 and not tok.isdigit():
            return tok
    return norm_sku.split()[0] if norm_sku.split() else ""


# ── Matching ──────────────────────────────────────────────────────────────────

def fuzzy_match(norm_sku, source_df):
    """
    Two-phase matching:
      Phase 1 — substring scan: filter rows where the product-family token
                appears in the short name, then score with token_sort_ratio
                against the cleaned short name (strips spec noise like
                ", w/csh DCZ...").  This correctly disambiguates items that
                share a product family but differ by type (corner vs ottoman,
                sidesofa vs 2-seater, etc.).
      Phase 2 — full fuzzy fallback: token_set_ratio across all short names.

    Returns (row_index, score) or (None, 0).
    """
    if source_df.empty:
        return None, 0

    short_norms = source_df["_short_norm"].tolist()
    family = _family_token(norm_sku)

    # Phase 1: narrow to family candidates, score on clean short names
    if len(family) >= 3:
        family_idx = [i for i, n in enumerate(short_norms) if family in n]
        if family_idx:
            best_i, best_score = None, 0
            for i in family_idx:
                s = fuzz.token_sort_ratio(norm_sku, short_norms[i])
                if s > best_score:
                    best_score, best_i = s, i
            if best_score >= SCORE_YELLOW:
                return best_i, best_score

    # Phase 2: full fuzzy scan on short names
    result = process.extractOne(norm_sku, short_norms, scorer=fuzz.token_set_ratio)
    if result is None:
        return None, 0
    _name, score, idx = result
    return idx, score


# ── Main pipeline ─────────────────────────────────────────────────────────────

def run():
    print(f"Loading target: {TARGET_FILE}")
    target_df = pd.read_excel(TARGET_FILE)
    target_df.columns = [str(c).strip() for c in target_df.columns]
    name_col = target_df.columns[0]
    items = [str(v).strip() for v in target_df[name_col].dropna().tolist()]
    print(f"  {len(items)} items to fill\n")

    print("Loading source files...")
    sources = load_all_sources()
    if sources.empty:
        print("ERROR: No source data loaded.")
        return
    print(f"  Total source records: {len(sources)}\n")

    print("Matching...")
    # Group B-suffix variants by their base SKU key
    base_groups = defaultdict(list)
    for sku in items:
        base = strip_box_suffix(sku)
        base_groups[base].append(sku)

    rows_out = []
    for base_sku, variants in base_groups.items():
        norm = normalize_sku(base_sku)
        idx, score = fuzzy_match(norm, sources)

        matched_name = ""
        L = W = H = gw_lbs = None
        source_label = ""

        if idx is not None:
            row = sources.iloc[idx]
            matched_name = row["item_name"]
            L, W, H      = row["L_cm"], row["W_cm"], row["H_cm"]
            gw_lbs       = row["GW_lbs"]
            source_label = row["source"]

        flag = "OK" if score >= SCORE_GREEN else ("??" if score >= SCORE_YELLOW else "--")
        print(f"  {flag} [{int(score):3d}] {base_sku[:38]:<38}  ->  {matched_name[:40]}")

        for i, sku in enumerate(variants):
            note = "copied" if len(variants) > 1 and i > 0 else ""
            def _cm_to_in(v):
                return round(v * CM_TO_IN, 2) if v is not None else None

            rows_out.append({
                "Name":         sku,
                "L_in":         _cm_to_in(L),
                "W_in":         _cm_to_in(W),
                "H_in":         _cm_to_in(H),
                "GW_lbs":       gw_lbs,
                "Matched_Item": matched_name,
                "Score":        score,
                "Source":       source_label,
                "Note":         note,
            })

    out_df = pd.DataFrame(rows_out)

    # Demote green rows with no usable dimension data → red
    has_dims = out_df[["L_in", "W_in", "H_in", "GW_lbs"]].notna().all(axis=1)
    no_dims_green = (out_df["Score"] >= SCORE_GREEN) & ~has_dims
    if no_dims_green.any():
        n_demoted = no_dims_green.sum()
        print(f"\n  WARNING: {n_demoted} green row(s) demoted to red (matched but no dimension data):")
        for name in out_df.loc[no_dims_green, "Name"].tolist():
            print(f"    {name}")
        out_df.loc[no_dims_green, "Note"] = out_df.loc[no_dims_green, "Note"].apply(
            lambda n: ("no dims in source; " + n).strip("; ") if n else "no dims in source"
        )
        out_df.loc[no_dims_green, "Score"] = 0  # force to red band for colouring/sorting

    # Sort: green first, then yellow, then red; score descending within each tier
    def _tier(score):
        if score >= SCORE_GREEN:   return 0
        if score >= SCORE_YELLOW:  return 1
        return 2
    out_df["_tier"] = out_df["Score"].apply(_tier)
    out_df = out_df.sort_values(["_tier", "Score"], ascending=[True, False]) \
                   .drop(columns=["_tier"]).reset_index(drop=True)

    # Stats
    n_green  = (out_df["Score"] >= SCORE_GREEN).sum()
    n_yellow = ((out_df["Score"] >= SCORE_YELLOW) & (out_df["Score"] < SCORE_GREEN)).sum()
    n_red    = (out_df["Score"] < SCORE_YELLOW).sum()
    print(f"\n  Green  (>={SCORE_GREEN}, auto-fill) : {n_green}")
    print(f"  Yellow ({SCORE_YELLOW}-{SCORE_GREEN-1}, review)   : {n_yellow}")
    print(f"  Red    (<{SCORE_YELLOW},  no match)  : {n_red}")

    # Write output
    print(f"\nWriting {OUTPUT_FILE}...")
    out_df.to_excel(OUTPUT_FILE, index=False, engine="openpyxl")

    # Colour-code rows by confidence
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")
    FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
    FILL_RED    = PatternFill("solid", fgColor="FFC7CE")

    for i, score in enumerate(out_df["Score"]):
        xl_row = i + 2   # 1-indexed + 1 for header
        fill   = FILL_GREEN if score >= SCORE_GREEN else (FILL_YELLOW if score >= SCORE_YELLOW else FILL_RED)
        for cell in ws[xl_row]:
            cell.fill = fill

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(OUTPUT_FILE)
    print("Done!")
    print(f"  Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    run()
