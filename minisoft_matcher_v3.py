"""
Minisoft Packing Match Logic v3 — Component-Level Decomposition

For kits missing Packing Method - Minisoft records, this script:
  1. Tries COMPONENT DECOMPOSITION first:
       - Classifies each kit component as table / chair / legs / frame / skip
       - Tables: look up the component's standalone box spec (exact dims & weight)
       - Chairs: look up empirical stacked-box spec (derived from pure-chair source kits)
         and emit ceil(qty / capacity) boxes
       - Legs / Frames: look up standalone box spec
  2. Falls back to EXACT match (same component set + qty → copy source boxes)
  3. Falls back to PARTIAL match (similarity score → proportional scale,
       component-aware for table+chair kits: table box fixed, chair boxes scaled)
  4. Falls back to NO MATCH if score < MIN_MATCH_SCORE

Matching is done only against Box-style source kits (custrecord5 IS NULL).
"""

import json
import math
import os
import statistics
import pandas as pd
from collections import Counter, defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── File paths ───────────────────────────────────────────────────────────────
FILE_WITH_MINISOFT = r'C:\Users\info\Documents\Dev\data\minisoft\with_minisoft.json'
FILE_ALL_MEMBERS   = r'C:\Users\info\Documents\Dev\data\minisoft\all_members.json'
FILE_NO_MINISOFT   = r'C:\Users\info\Documents\Dev\data\minisoft\no_minisoft.json'
FILE_CROSS_REF     = r'C:\Users\info\Documents\Dev\data\minisoft\cross_ref.json'
FILE_COMP_DESC     = r'C:\Users\info\Documents\Dev\data\minisoft\comp_descriptions.json'
OUTPUT_FILE        = os.environ.get(
    'MINISOFT_OUTPUT_FILE',
    r'C:\Users\info\Documents\Dev\Output\Minisoft_Match_v3.xlsx'
)

EXCLUDED_MATCH_KIT_NAMES = {'SC AMA KIT_OIL NEW'}

# ── Component role keywords (name-based fallback) ─────────────────────────────
# Order matters: skip checked first, then table before chair.
# BT prefix removed from TABLE_KW — many BT items are chairs (use description instead).
_SKIP_KW  = ('MAINTKIT', 'KIT_OIL', 'CUSHION', 'COVER', 'PARASOL', 'UMBRELLA')
_TABLE_KW = ('PLI ', 'SC ALAMA', 'SC CHAMONIX_RECT', 'SC CANNES_160',
             'SC CANNES_TABLE', 'SC DIAN', 'SC LEYLAND', '_RECT', ' TABLE',
             'GC_ATL', 'IN_CTECHA', 'BT 5', 'BT 3', 'BT 4', 'BT 2')  # BT numeric = table
_CHAIR_KW = ('SIDE', 'BUCKET', 'CHAIR', 'CHAISE', 'BARIARMC', 'BARI_ARM',
             'CANNES_LOT', 'CANNES_SIDE', 'CANNES_BUCKET')
_SOFA_KW  = ('SOFA', '2SEATER', '3SEATER', '2 SEATER', '3 SEATER')
_SECT_KW  = ('OTTOMAN', 'LOUNGER', 'CHAISE_LONG', 'SECTIONAL')
_LEGS_KW  = ('_LEGS_', ' LEGS ', 'LEGS_')
_FRAME_KW = ('_FRAME_', ' FRAME_', '_FRAME ')

# Description-based keywords (checked before name keywords when description exists)
_DESC_SKIP = ('cushion', 'cover', 'parasol', 'umbrella', 'screen', 'chest')
_DESC_TABLE = ('table', 'bartable', 'bar table', 'counter height')
_DESC_SECT  = ('sectional', 'loveseat', 'chaise lounge', 'ottoman')
_DESC_SOFA  = ('sofa', '2-seater', '3-seater', '2 seater', '3 seater')
_DESC_CHAIR = ('chair', 'barstool', 'bar stool', 'stool', 'deckchair',
               'deck chair', 'armchair', 'bench')

DEFAULT_CHAIR_CAPACITY = 4   # empirically confirmed: 4 chairs stacked per box

MIN_MATCH_SCORE = 0.40

# ── Component family matching ─────────────────────────────────────────────────
# Tokens stripped when building a chair's "family key" for fuzzy matching.
# Color/finish variants of the same model share the same family key.
import re as _re
_FAMILY_COLOR   = {'WHT','WH','GRY','GR','GRN','BLK','BK','HOY','PNY','SGN',
                   'BMB','MGR','OR','DGRY','CSNM','NTP','ZWHT','BEI','T','RCY',
                   'OW','SG','DGR','NAT','NOK','BN','CRM'}
_FAMILY_FINISH  = {'LOT','PAR','ALU','FSC','TEA','STL','RSN','LOTS','PARS','ALUS'}
_FAMILY_SUBMDL  = {'SIDE','ARM','BUCKET','CARVER','DINING','STACKING','FOLDING',
                   'LOUNGE','BARIARMC','CHAIR','ARMCHAIR'}
_FAMILY_STRIP   = _FAMILY_COLOR | _FAMILY_FINISH | _FAMILY_SUBMDL

def comp_family(cid: str) -> str:
    """Return a normalized family key for a component, stripping color/finish/sub-model tokens."""
    name = comp_name_map.get(cid, str(cid))
    tokens = _re.split(r'[\s_\-]+', name.strip().upper())
    kept = [t for t in tokens if t and t not in _FAMILY_STRIP]
    return '_'.join(kept) if kept else name.upper()


# ── Helpers ──────────────────────────────────────────────────────────────────

def load_json(path):
    with open(path, encoding='utf-8') as f:
        raw = f.read()
    outer = json.loads(raw)
    # Support both MCP tool-result format and plain JSON array
    if isinstance(outer, list) and outer and isinstance(outer[0], dict) and 'text' in outer[0]:
        return json.loads(outer[0]['text'])['data']
    return outer


def get_pkg_type(record):
    return 'Pallet' if record.get('num_boxes') is not None else 'Box'


def _has_box_payload(record):
    """True when a Box-style record has usable dimensional/weight payload."""
    if record.get('num_boxes') is not None:
        return False
    return any(record.get(k) is not None for k in ('weight', 'length', 'width', 'height'))


def coerce_qty(value):
    """Normalize quantity values from JSON/API payloads to numeric."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return value
    try:
        v = float(str(value).strip())
    except (TypeError, ValueError):
        return 0
    return int(v) if v.is_integer() else v


def classify_component(name: str, desc: str = '') -> str:
    """Return 'table' | 'chair' | 'sofa' | 'sectional' | 'legs' | 'frame' | 'skip' | 'other'.

    Description (when available) takes priority over item-name keywords.
    """
    n = name.upper()
    d = desc.lower() if desc else ''

    # Skip is checked first regardless of source
    if any(k.upper() in n for k in _SKIP_KW):
        return 'skip'
    if d and any(k in d for k in _DESC_SKIP):
        return 'skip'

    # Description-based classification (more reliable)
    if d:
        if any(k in d for k in _DESC_SECT):
            return 'sectional'
        if any(k in d for k in _DESC_SOFA):
            return 'sofa'
        if any(k in d for k in _DESC_CHAIR):
            return 'chair'
        if any(k in d for k in _DESC_TABLE):
            return 'table'

    # Name-based fallback
    if any(k.upper() in n for k in _SECT_KW):
        return 'sectional'
    if any(k.upper() in n for k in _SOFA_KW):
        return 'sofa'
    if any(k.upper() in n for k in _TABLE_KW):
        return 'table'
    if any(k.upper() in n for k in _CHAIR_KW):
        return 'chair'
    if any(k.upper() in n for k in _LEGS_KW):
        return 'legs'
    if any(k.upper() in n for k in _FRAME_KW):
        return 'frame'
    return 'other'


def classify_cid(cid: str) -> str:
    """Classify a component by ID, using description when available."""
    name = comp_name_map.get(cid, str(cid))
    desc = comp_desc_map.get(str(cid), '')
    return classify_component(name, desc)


def to_qty_dict(qty_set):
    return {item: qty for item, qty in qty_set}


def is_excluded_source_kit(kit_id):
    return kit_name_map.get(kit_id, '').strip().upper() in {n.upper() for n in EXCLUDED_MATCH_KIT_NAMES}


def get_kit_pack_type(kit_id):
    recs = minisoft_by_kit.get(kit_id, [])
    has_pallet = any(r.get('num_boxes') is not None for r in recs)
    has_box_payload = any(_has_box_payload(r) for r in recs)

    # Prefer pallet when box records are placeholders with no usable payload.
    if has_box_payload and not has_pallet:
        return 'Box'
    if has_pallet and not has_box_payload:
        return 'Pallet'
    if has_box_payload and has_pallet:
        return 'Box'
    return None


# ── Load raw data ─────────────────────────────────────────────────────────────
print("Loading data...")
rows_with    = load_json(FILE_WITH_MINISOFT)
rows_all     = load_json(FILE_ALL_MEMBERS)
rows_missing = load_json(FILE_NO_MINISOFT)
rows_xref    = load_json(FILE_CROSS_REF)

# Component descriptions {item_id_str: {itemid, description}}
with open(FILE_COMP_DESC, encoding='utf-8') as _f:
    _comp_desc_raw = json.load(_f)
comp_desc_map = {cid: v.get('description', '') for cid, v in _comp_desc_raw.items()}

xref_kit_ids = {r['item_id'] for r in rows_xref}
print(f"Cross-referenced kits: {len(xref_kit_ids)}")

# ── Missing kits ──────────────────────────────────────────────────────────────
missing_kits = {
    r['id']: r['itemid']
    for r in rows_missing
    if r['itemtype'] in ('Kit', 'Package') and r['id'] in xref_kit_ids
}
missing_kit_ids = set(missing_kits.keys())
print(f"Kits missing Minisoft (cross-referenced): {len(missing_kit_ids)}")

# ── All kit members ───────────────────────────────────────────────────────────
kit_components_ids   = defaultdict(set)
kit_components_qty   = defaultdict(set)
kit_name_map         = {}
comp_name_map        = {}

for row in rows_all:
    pid = row['parentitem']
    cid = row['component_item']
    qty = coerce_qty(row['quantity'])
    kit_components_ids[pid].add(cid)
    kit_components_qty[pid].add((cid, qty))
    kit_name_map[pid] = row['kit_name']
    comp_name_map[cid] = row['component_name']

kit_components_ids = {k: frozenset(v) for k, v in kit_components_ids.items()}
kit_components_qty = {k: frozenset(v) for k, v in kit_components_qty.items()}


def comp_label(kit_id):
    pairs = sorted(
        (comp_name_map.get(cid, str(cid)), qty)
        for cid, qty in kit_components_qty.get(kit_id, set())
    )
    return ', '.join(f'{n} x{q}' for n, q in pairs)


# ── Minisoft records for source kits ──────────────────────────────────────────
seen_minisoft         = {}
kit_minisoft_records  = defaultdict(list)
kit_with_minisoft_ids = defaultdict(set)
kit_with_minisoft_qty = defaultdict(set)

for row in rows_with:
    pid = row['parentitem']
    mid = row['minisoft_id']
    cid = row['component_item']
    qty = coerce_qty(row['quantity'])
    kit_with_minisoft_ids[pid].add(cid)
    kit_with_minisoft_qty[pid].add((cid, qty))
    kit_name_map[pid] = row['kit_name']
    comp_name_map[cid] = row['component_name']
    if mid not in seen_minisoft:
        seen_minisoft[mid] = {
            'minisoft_id' : mid,
            'created'     : row['created'],
            'pkg_number'  : row['custrecord4'],
            'num_boxes'   : row['custrecord5'],
            'weight'      : row['custrecord6'],
            'length'      : row['custrecord7'],
            'width'       : row['custrecord8'],
            'height'      : row['custrecord9'],
        }
        kit_minisoft_records[pid].append(seen_minisoft[mid])

kit_with_minisoft_ids = {k: frozenset(v) for k, v in kit_with_minisoft_ids.items()}
kit_with_minisoft_qty = {k: frozenset(v) for k, v in kit_with_minisoft_qty.items()}

minisoft_by_kit = {
    kid: sorted(recs, key=lambda r: (float(r['pkg_number']) if r['pkg_number'] is not None else 0,
                                     r['minisoft_id']))
    for kid, recs in kit_minisoft_records.items()
}
kits_with_minisoft = set(minisoft_by_kit.keys())
print(f"Kits WITH Minisoft: {len(kits_with_minisoft)}")


# ── Component Box Library ─────────────────────────────────────────────────────
def build_component_box_library():
    """
    Scan source kits for single-component (qty=1) Box-style kits.
    Those box records are the per-unit standalone spec for that component.
    Returns {comp_id: [sorted box records]}.
    """
    library = {}
    for kit_id in kits_with_minisoft:
        if is_excluded_source_kit(kit_id):
            continue
        box_recs = [r for r in minisoft_by_kit[kit_id] if r.get('num_boxes') is None]
        if not box_recs:
            continue
        comps = kit_with_minisoft_qty.get(kit_id, frozenset())
        non_skip = [
            (cid, int(qty)) for cid, qty in comps
            if classify_cid(cid) != 'skip'
        ]
        if len(non_skip) == 1 and non_skip[0][1] == 1:
            comp_id = non_skip[0][0]
            if comp_id not in library:
                library[comp_id] = sorted(
                    box_recs,
                    key=lambda r: float(r['pkg_number']) if r['pkg_number'] is not None else 0
                )
    return library


# ── Chair Spec Library ────────────────────────────────────────────────────────
def build_chair_spec_library():
    """
    Scan source kits that have ONLY chair components (no table).
    Derive the empirical stacked-box spec per chair model:
      - capacity (chairs per box)
      - weight_per_unit (lbs per chair when stacked)
      - box dims (L, W, H) averaged across source kits
    Returns {chair_comp_id: spec_dict}.
    """
    data = defaultdict(list)   # chair_id → [(qty, [box_records])]

    for kit_id in kits_with_minisoft:
        if is_excluded_source_kit(kit_id):
            continue
        box_recs = [r for r in minisoft_by_kit[kit_id] if r.get('num_boxes') is None]
        if not box_recs:
            continue
        comps = kit_with_minisoft_qty.get(kit_id, frozenset())
        roles = {cid: classify_cid(cid) for cid, _ in comps}

        # Pure-chair source only: ignore skip items, then require all remaining roles to be chair.
        non_skip_roles = [r for r in roles.values() if r != 'skip']
        if not non_skip_roles:
            continue
        if any(r != 'chair' for r in non_skip_roles):
            continue

        # Must have exactly one chair component type
        chair_comps = [(cid, int(qty)) for cid, qty in comps if roles[cid] == 'chair']
        if len(chair_comps) != 1:
            continue

        chair_id, chair_qty = chair_comps[0]
        if chair_qty < 1:
            continue
        data[chair_id].append((chair_qty, box_recs))

    library = {}
    for chair_id, obs in data.items():
        capacities, wpus, dims_list = [], [], []
        for qty, boxes in obs:
            n = len(boxes)
            if n == 0 or qty == 0:
                continue
            cap = qty / n
            if 0.5 <= cap <= 12:
                capacities.append(cap)
            weights = [float(r['weight']) for r in boxes if r.get('weight') is not None]
            if weights:
                wpus.append(sum(weights) / qty)
            sorted_b = sorted(boxes, key=lambda r: float(r['pkg_number']) if r['pkg_number'] is not None else 0)
            b = sorted_b[0]
            try:
                dims_list.append((float(b['length']), float(b['width']), float(b['height'])))
            except (TypeError, ValueError):
                pass

        if not capacities:
            continue
        cap = round(statistics.mean(capacities))
        wpu = round(statistics.mean(wpus), 2) if wpus else None
        if dims_list:
            avg_l = round(statistics.mean(d[0] for d in dims_list), 1)
            avg_w = round(statistics.mean(d[1] for d in dims_list), 1)
            avg_h = round(statistics.mean(d[2] for d in dims_list), 1)
        else:
            avg_l = avg_w = avg_h = None

        library[chair_id] = {
            'capacity'       : cap,
            'weight_per_unit': wpu,
            'length'         : avg_l,
            'width'          : avg_w,
            'height'         : avg_h,
            'obs_count'      : len(obs),
        }

    return library


component_box_library = build_component_box_library()
chair_spec_library    = build_chair_spec_library()
print(f"Component box library: {len(component_box_library)} items with standalone specs")
print(f"Chair spec library: {len(chair_spec_library)} chair models with stacked specs")


# ── Component Decomposition ───────────────────────────────────────────────────
def infer_boxes_from_components(target_kit_id):
    """
    Decompose target kit into per-component boxes using the libraries.
    Returns a list of box dicts (one per box) with source_component / source_type,
    or None if any required spec is missing.
    """
    comps = kit_components_qty.get(target_kit_id, frozenset())
    if not comps:
        return None

    output_boxes = []
    box_num = 1
    unresolved = []

    # Sort by role priority: tables first, then chairs, then legs/frames
    role_order = {'table': 0, 'chair': 1, 'sofa': 2, 'sectional': 3, 'legs': 4, 'frame': 5, 'other': 6, 'skip': 7}
    sorted_comps = sorted(
        comps,
        key=lambda x: (role_order.get(classify_cid(x[0]), 9),
                       comp_name_map.get(x[0], ''))
    )

    for comp_id, qty in sorted_comps:
        qty = int(qty)
        name = comp_name_map.get(comp_id, str(comp_id))
        role = classify_cid(comp_id)

        if role == 'skip':
            continue

        if role in ('sofa', 'sectional'):
            # Sofas/sectionals each ship individually — use standalone spec per unit
            spec = component_box_library.get(comp_id)
            if spec is None:
                unresolved.append(f'{name} ({role}, no standalone spec)')
                continue
            for _ in range(qty):
                for b in spec:
                    output_boxes.append({
                        'pkg_number'     : box_num,
                        'weight'         : b.get('weight'),
                        'length'         : b.get('length'),
                        'width'          : b.get('width'),
                        'height'         : b.get('height'),
                        'source_component': name,
                        'source_type'    : f'component:{role}',
                    })
                    box_num += 1
            continue

        if role == 'table':
            spec = component_box_library.get(comp_id)
            if spec is None:
                unresolved.append(f'{name} (table, no standalone spec)')
                continue
            for _ in range(qty):   # usually qty=1 for tables
                for b in spec:
                    output_boxes.append({
                        'pkg_number'     : box_num,
                        'weight'         : b.get('weight'),
                        'length'         : b.get('length'),
                        'width'          : b.get('width'),
                        'height'         : b.get('height'),
                        'source_component': name,
                        'source_type'    : 'component:table',
                    })
                    box_num += 1

        elif role == 'chair':
            spec = chair_spec_library.get(comp_id)
            if spec is None:
                # Try standalone spec as fallback (single chair packing)
                standalone = component_box_library.get(comp_id)
                if standalone:
                    cap = DEFAULT_CHAIR_CAPACITY
                    # Estimate stacked weight from standalone (rough: stacked ≈ standalone)
                    w_each = float(standalone[0]['weight']) if standalone[0].get('weight') else None
                    for _ in range(math.ceil(qty / cap)):
                        n_in = min(cap, qty - (_ * cap))
                        if n_in < 1:
                            break
                        w = round(n_in * w_each, 1) if w_each else None
                        b = standalone[0]
                        output_boxes.append({
                            'pkg_number'     : box_num,
                            'weight'         : str(w) if w else None,
                            'length'         : b.get('length'),
                            'width'          : b.get('width'),
                            'height'         : b.get('height'),
                            'source_component': name,
                            'source_type'    : f'component:chair(standalone_fallback,cap={cap})',
                        })
                        box_num += 1
                else:
                    unresolved.append(f'{name} (chair, no spec)')
                continue
            cap = spec['capacity'] or DEFAULT_CHAIR_CAPACITY
            wpu = spec['weight_per_unit']
            remaining = qty
            while remaining > 0:
                n_in = min(cap, remaining)
                w = round(n_in * wpu, 1) if wpu is not None else None
                output_boxes.append({
                    'pkg_number'     : box_num,
                    'weight'         : str(w) if w is not None else None,
                    'length'         : str(spec['length']) if spec['length'] else None,
                    'width'          : str(spec['width']) if spec['width'] else None,
                    'height'         : str(spec['height']) if spec['height'] else None,
                    'source_component': name,
                    'source_type'    : f'component:chair(stacked,{n_in}/{cap}perbox,{spec["obs_count"]}kits)',
                })
                remaining -= n_in
                box_num += 1

        else:   # legs, frame, other → use standalone spec if available
            spec = component_box_library.get(comp_id)
            if spec is None:
                unresolved.append(f'{name} ({role}, no standalone spec)')
                continue
            for _ in range(qty):
                for b in spec:
                    output_boxes.append({
                        'pkg_number'     : box_num,
                        'weight'         : b.get('weight'),
                        'length'         : b.get('length'),
                        'width'          : b.get('width'),
                        'height'         : b.get('height'),
                        'source_component': name,
                        'source_type'    : f'component:{role}',
                    })
                    box_num += 1

    if unresolved:
        return None   # Incomplete decomposition → fall back to similarity matching
    if not output_boxes:
        return None
    return output_boxes


# ── Similarity Matching (existing logic) ─────────────────────────────────────
qty_set_to_kits = defaultdict(list)
for kid in kits_with_minisoft:
    if is_excluded_source_kit(kid):
        continue
    qset = kit_with_minisoft_qty.get(kid, frozenset())
    if qset:
        qty_set_to_kits[qset].append(kid)

source_qty_dicts = {
    kid: to_qty_dict(kit_with_minisoft_qty.get(kid, frozenset()))
    for kid in kits_with_minisoft
    if not is_excluded_source_kit(kid)
}

peer_index = defaultdict(list)
for kid in kits_with_minisoft:
    if is_excluded_source_kit(kid):
        continue
    ids_key = kit_with_minisoft_ids.get(kid, frozenset())
    if ids_key:
        peer_index[ids_key].append(kid)
print(f"Peer groups: {len(peer_index)}")

# Indexes for seating match narrowing.
# 1) seating_component_id -> source kits (exact component match)
# 2) seating_family_key   -> source kits (family fallback)
_SEATING_ROLES = {'chair', 'sofa', 'sectional'}
chair_to_source_kits = defaultdict(set)
family_to_source_kits = defaultdict(set)
for kid in kits_with_minisoft:
    if is_excluded_source_kit(kid):
        continue
    for cid in kit_with_minisoft_ids.get(kid, frozenset()):
        if classify_cid(cid) in _SEATING_ROLES:
            chair_to_source_kits[cid].add(kid)
            family_to_source_kits[comp_family(cid)].add(kid)


def similarity_score(target_dict, source_dict):
    shared = set(target_dict) & set(source_dict)
    if not shared:
        return 0.0
    target_recall    = len(shared) / len(target_dict)
    source_precision = len(shared) / len(source_dict)
    qty_sims = []
    for i in shared:
        t = coerce_qty(target_dict[i])
        s = coerce_qty(source_dict[i])
        denom = max(t, s)
        if denom > 0:
            qty_sims.append(min(t, s) / denom)
    qty_component = (sum(qty_sims) / len(qty_sims)) if qty_sims else 0.0
    return target_recall * 0.45 + source_precision * 0.25 + qty_component * 0.30


def get_kit_totals(kit_id):
    recs        = minisoft_by_kit.get(kit_id, [])
    box_recs    = [r for r in recs if _has_box_payload(r)]
    pallet_recs = [r for r in recs if r['num_boxes'] is not None]
    if box_recs:
        return (sum(float(r['weight']) for r in box_recs if r['weight'] is not None),
                len(box_recs))
    if pallet_recs:
        return (sum(float(r['weight']) for r in pallet_recs if r['weight'] is not None),
                sum(int(r['num_boxes']) for r in pallet_recs))
    return (0, 0)


def linear_predict(x_vals, y_vals, x_target):
    n = len(x_vals)
    if n < 2:
        return y_vals[0] * x_target / x_vals[0] if x_vals[0] else y_vals[0]
    x_mean = sum(x_vals) / n
    y_mean = sum(y_vals) / n
    num = sum((x_vals[i] - x_mean) * (y_vals[i] - y_mean) for i in range(n))
    den = sum((x_vals[i] - x_mean) ** 2 for i in range(n))
    if den == 0:
        return y_mean
    slope = num / den
    return slope * x_target + (y_mean - slope * x_mean)


def _pkg_sort_key(record):
    pkg = record.get('pkg_number')
    try:
        return (float(pkg) if pkg is not None else float('inf'), record.get('minisoft_id', 0))
    except (TypeError, ValueError):
        return (float('inf'), record.get('minisoft_id', 0))


def _filter_chair_box_outliers(records):
    """
    Remove obvious non-chair outlier boxes (e.g., tiny accessory boxes) from a
    table+chair source kit before chair-box scaling.
    """
    parsed = []
    for r in records:
        try:
            l = float(r.get('length') or 0)
            w = float(r.get('width') or 0)
            h = float(r.get('height') or 0)
        except (TypeError, ValueError):
            continue
        if l > 0 and w > 0 and h > 0:
            parsed.append((r, l, w, h))

    if len(parsed) < 3:
        return records

    med_fp = statistics.median(l * w for _, l, w, _ in parsed)
    med_h  = statistics.median(h for _, _, _, h in parsed)
    kept = [
        r for r, l, w, h in parsed
        if (l * w) >= (0.5 * med_fp) and h >= (0.5 * med_h)
    ]
    return kept if kept else records


def identify_table_box(matched_id):
    """
    For a Box-style kit, identify which box record is the table box.
    Heuristic: the box with the largest footprint (L × W) and small H is the table.
    Returns the index (0-based) of the table box among sorted box records, or None.
    """
    box_recs = sorted(
        [r for r in minisoft_by_kit.get(matched_id, []) if r.get('num_boxes') is None],
        key=_pkg_sort_key
    )
    if not box_recs:
        return None, []

    # Score each box: table boxes have large L*W and small H
    def table_score(r):
        try:
            l, w, h = float(r.get('length') or 0), float(r.get('width') or 0), float(r.get('height') or 0)
            if h == 0:
                return 0
            return (l * w) / h
        except (TypeError, ValueError):
            return 0

    # Check if any source components are tables
    source_comps = kit_with_minisoft_qty.get(matched_id, frozenset())
    has_table = any(
        classify_cid(cid) == 'table'
        for cid, _ in source_comps
    )
    if not has_table:
        return None, box_recs

    # Return index of highest table_score
    scores = [table_score(r) for r in box_recs]
    best_idx = scores.index(max(scores))
    return best_idx, box_recs


def infer_packing(target_kit_id, matched_kit_id):
    """
    Infer packing for target kit from matched kit.
    Component-aware for Box-style table+chair kits: table box is fixed,
    chair boxes are scaled by chair quantity ratio.
    """
    if not matched_kit_id:
        return None

    target_dict  = to_qty_dict(kit_components_qty.get(target_kit_id, frozenset()))
    matched_dict = source_qty_dicts.get(matched_kit_id, {})
    if not target_dict or not matched_dict:
        return None

    matched_pack_type = get_kit_pack_type(matched_kit_id)

    if matched_pack_type == 'Box':
        matched_weight, matched_boxes = get_kit_totals(matched_kit_id)
        if matched_boxes <= 0:
            return None

        # Determine variable (chair) component
        variable_item = None
        max_diff = 0
        for item in target_dict:
            if item in matched_dict:
                diff = abs(target_dict[item] - matched_dict[item])
                if diff > max_diff:
                    max_diff = diff
                    variable_item = item
        if variable_item is None:
            shared = set(target_dict) & set(matched_dict)
            if shared:
                variable_item = max(shared, key=lambda x: target_dict[x])
            else:
                return None

        target_x  = target_dict[variable_item]
        matched_x = matched_dict.get(variable_item, 0)
        if matched_x == 0 or target_x == 0:
            return None
        var_name = comp_name_map.get(variable_item, str(variable_item))
        var_role = classify_cid(variable_item)

        # Check if both kits have a table component
        target_has_table  = any(
            classify_cid(cid) == 'table'
            for cid in target_dict
        )
        matched_has_table = any(
            classify_cid(cid) == 'table'
            for cid in matched_dict
        )

        if var_role == 'chair' and target_has_table and matched_has_table:
            # Component-aware: fix the table box, scale only the chair boxes
            table_idx, box_recs = identify_table_box(matched_id=matched_kit_id)
            if table_idx is not None and len(box_recs) > 1:
                chair_boxes_all = [r for i, r in enumerate(box_recs) if i != table_idx]
                chair_boxes_source = _filter_chair_box_outliers(chair_boxes_all)
                matched_chair_boxes = len(chair_boxes_source)
                chair_weight_source = sum(
                    float(r['weight']) for r in chair_boxes_source if r.get('weight') is not None
                )
                ratio = target_x / matched_x
                pred_chair_boxes  = matched_chair_boxes * ratio
                pred_chair_weight = chair_weight_source * ratio
                inferred_boxes    = 1 + max(1, math.ceil(pred_chair_boxes))  # +1 for table
                inferred_weight   = float(box_recs[table_idx].get('weight') or 0) + pred_chair_weight
                dropped = len(chair_boxes_all) - len(chair_boxes_source)
                notes = (
                    f'Component-aware: table box fixed (pkg #{box_recs[table_idx]["pkg_number"]}), '
                    f'"{var_name}": {matched_x:.0f}→{target_x:.0f} units, '
                    f'{matched_chair_boxes:.0f}→{pred_chair_boxes:.1f} chair boxes'
                )
                if dropped > 0:
                    notes += f', excluded {dropped} outlier box(es)'
                return {
                    'inferred_weight': round(inferred_weight, 1) if inferred_weight > 0 else None,
                    'inferred_boxes' : inferred_boxes,
                    'inference_notes': notes,
                }

        # Fallback: pure proportional scaling for Box-style
        ratio = target_x / matched_x
        pred_boxes  = matched_boxes * ratio
        pred_weight = matched_weight * ratio if matched_weight > 0 else 0
        per_unit = matched_x / matched_boxes if matched_boxes > 0 else 0
        notes = (
            f'Box proportional: "{var_name}" — {matched_x:.0f} units in {matched_boxes:.0f} boxes '
            f'({per_unit:.1f}/box) → {target_x} units = {pred_boxes:.1f} boxes'
        )
        return {
            'inferred_weight': round(pred_weight, 1) if pred_weight > 0 else None,
            'inferred_boxes' : max(1, math.ceil(pred_boxes)),
            'inference_notes': notes,
        }

    # ── Pallet / regression path ──────────────────────────────────────────────
    variable_item = None
    max_diff = 0
    for item in target_dict:
        if item in matched_dict:
            diff = abs(target_dict[item] - matched_dict[item])
            if diff > max_diff:
                max_diff = diff
                variable_item = item
    if variable_item is None:
        shared = set(target_dict) & set(matched_dict)
        if shared:
            variable_item = max(shared, key=lambda x: target_dict[x])
        else:
            return None

    target_x  = target_dict[variable_item]
    matched_x = matched_dict.get(variable_item, 0)
    if matched_x == 0 or target_x == 0:
        return None
    var_name = comp_name_map.get(variable_item, str(variable_item))

    target_ids = kit_components_ids.get(target_kit_id, frozenset())
    points = []
    for peer_id in peer_index.get(target_ids, []):
        if matched_pack_type and get_kit_pack_type(peer_id) != matched_pack_type:
            continue
        pdict = source_qty_dicts.get(peer_id, {})
        if variable_item not in pdict:
            continue
        weight, boxes = get_kit_totals(peer_id)
        if weight > 0 or boxes > 0:
            points.append((pdict[variable_item], weight, boxes))

    if len(points) >= 2:
        points.sort(key=lambda p: p[0])
        x_vals = [p[0] for p in points]
        pred_weight = linear_predict(x_vals, [p[1] for p in points], target_x)
        pred_boxes  = linear_predict(x_vals, [p[2] for p in points], target_x)
        pts_str = '; '.join(f'qty={px}: {pw:.0f}lbs,{pb:.0f}boxes' for px, pw, pb in points)
        notes = f'Regression on "{var_name}" ({len(points)} peers): [{pts_str}] → qty={target_x}'
    else:
        matched_weight, matched_boxes = get_kit_totals(matched_kit_id)
        if matched_weight == 0 and matched_boxes == 0:
            return None
        ratio = target_x / matched_x
        pred_boxes  = matched_boxes * ratio
        pred_weight = matched_weight * ratio
        per_unit = matched_x / matched_boxes if matched_boxes > 0 else 0
        notes = (
            f'Proportional: "{var_name}" — {matched_x:.0f} units in {matched_boxes:.0f} boxes '
            f'({per_unit:.1f}/box) → {target_x} units = {pred_boxes:.1f} boxes'
        )

    return {
        'inferred_weight': round(pred_weight, 1) if pred_weight > 0 else None,
        'inferred_boxes' : max(0, round(pred_boxes)),
        'inference_notes': notes,
    }


def select_output_records(matched_id, packing_basis, inferred_boxes, match_type):
    """Pick source Minisoft rows to emit for a similarity-matched kit."""
    recs = list(minisoft_by_kit.get(matched_id, []))
    if not recs:
        return recs

    if packing_basis == 'Box':
        recs = [r for r in recs if r.get('num_boxes') is None]
    elif packing_basis == 'Pallet':
        pallet_recs = [r for r in recs if r.get('num_boxes') is not None]
        recs = pallet_recs if pallet_recs else recs

    recs = sorted(recs, key=_pkg_sort_key)

    if packing_basis == 'Box' and str(match_type).startswith('Partial'):
        try:
            target_boxes = int(inferred_boxes)
        except (TypeError, ValueError):
            target_boxes = 0
        if target_boxes > 0:
            by_pkg = {}
            for r in recs:
                pkg = r.get('pkg_number')
                if pkg is None:
                    continue
                try:
                    key = int(float(pkg))
                except (TypeError, ValueError):
                    continue
                if key not in by_pkg:
                    by_pkg[key] = r
            if by_pkg:
                chosen = [by_pkg[p] for p in sorted(by_pkg) if p <= target_boxes]
                if len(chosen) < target_boxes:
                    used_ids = {r.get('minisoft_id') for r in chosen}
                    for r in recs:
                        if r.get('minisoft_id') in used_ids:
                            continue
                        chosen.append(r)
                        used_ids.add(r.get('minisoft_id'))
                        if len(chosen) >= target_boxes:
                            break
                if len(chosen) < target_boxes:
                    # Expand with repeated chair-box templates when source has fewer
                    # rows than predicted target boxes.
                    table_idx, box_recs = identify_table_box(matched_id)
                    if table_idx is not None and box_recs:
                        filler_pool = [r for i, r in enumerate(box_recs) if i != table_idx]
                    else:
                        filler_pool = []
                    if not filler_pool:
                        filler_pool = list(chosen) if chosen else list(recs)
                    i = 0
                    while len(chosen) < target_boxes and filler_pool:
                        chosen.append(dict(filler_pool[i % len(filler_pool)]))
                        i += 1
                recs = chosen[:target_boxes] if chosen else recs[:target_boxes]
            else:
                recs = recs[:target_boxes]

    return recs


# ── Build base row template ────────────────────────────────────────────────────
def _empty_row(kit_id, kit_name, my_label, match_type):
    return {
        'Kit ID'             : kit_id,
        'Kit Name'           : kit_name,
        'My Components'      : my_label,
        'Match Type'         : match_type,
        'Match Score'        : '',
        'Matched Kit ID'     : '',
        'Matched Kit Name'   : '',
        'Source Components'  : '',
        'Date Created'       : '',
        'Package Number'     : '',
        'Package Type'       : 'Box',
        'Num Boxes in Pallet': '',
        'Weight'             : '',
        'Length'             : '',
        'Width'              : '',
        'Height'             : '',
        'Source Component'   : '',
        'Source Type'        : '',
        'Inferred Boxes'     : '',
        'Total Inferred Boxes': '',
        'Inferred Weight'    : '',
        'Packing Basis'      : '',
        'Inference Notes'    : '',
    }


# ── Main matching loop ────────────────────────────────────────────────────────
print("\nMatching kits...")
results = []

for kit_id in sorted(missing_kit_ids):
    kit_name = missing_kits[kit_id]
    my_ids   = kit_components_ids.get(kit_id, frozenset())
    my_qty   = kit_components_qty.get(kit_id, frozenset())
    my_label = comp_label(kit_id)

    if not my_ids:
        row = _empty_row(kit_id, kit_name, '', 'No components found')
        results.append(row)
        continue

    my_dict = to_qty_dict(my_qty)

    # ── Phase 1: Component Decomposition ──────────────────────────────────────
    decomp_boxes = infer_boxes_from_components(kit_id)
    if decomp_boxes:
        for b in decomp_boxes:
            row = _empty_row(kit_id, kit_name, my_label, 'Component Decomposition')
            row['Package Number']   = b['pkg_number']
            row['Weight']           = b.get('weight', '')
            row['Length']           = b.get('length', '')
            row['Width']            = b.get('width', '')
            row['Height']           = b.get('height', '')
            row['Source Component'] = b.get('source_component', '')
            row['Source Type']      = b.get('source_type', '')
            row['Inferred Boxes']   = len(decomp_boxes)
            row['Packing Basis']    = 'Box'
            results.append(row)
        continue   # Don't fall through to similarity matching

    # ── Phase 2: Exact match ──────────────────────────────────────────────────
    exact = qty_set_to_kits.get(my_qty, [])
    if exact:
        match_type = 'Exact'
        matched_id = exact[0]
        best_score = 1.0
    else:
        # ── Phase 3: Partial similarity match ─────────────────────────────────
        # If the target has seating components, prioritize exact component overlap.
        # If no exact overlap exists, fall back to seating family overlap.
        target_seating_ids = frozenset(
            cid for cid in my_ids
            if classify_cid(cid) in _SEATING_ROLES
        )
        if target_seating_ids:
            allowed_sources = frozenset(
                kid for cid in target_seating_ids
                for kid in chair_to_source_kits.get(cid, set())
            )
            if not allowed_sources:
                target_seating_families = frozenset(comp_family(cid) for cid in target_seating_ids)
                allowed_sources = frozenset(
                    kid for fam in target_seating_families
                    for kid in family_to_source_kits.get(fam, set())
                )
            if not allowed_sources:
                allowed_sources = None  # no seating candidates found; avoid false "No match"
        else:
            allowed_sources = None  # no restriction
        target_has_table = any(classify_cid(cid) == 'table' for cid in my_ids)
        target_has_seating = bool(target_seating_ids)

        best_score = 0.0
        matched_id = None
        for kid, src_dict in source_qty_dicts.items():
            if allowed_sources is not None and kid not in allowed_sources:
                continue
            if target_has_table and target_has_seating and get_kit_pack_type(kid) != 'Box':
                continue
            score = similarity_score(my_dict, src_dict)
            if score > best_score:
                best_score = score
                matched_id = kid
        if matched_id and best_score >= MIN_MATCH_SCORE:
            match_type = 'Partial (similar components)'
        else:
            match_type = 'No match'
            matched_id = None

    # Infer packing totals for partial matches
    inference = None
    if match_type.startswith('Partial'):
        inference = infer_packing(kit_id, matched_id)

    inferred_boxes  = inference['inferred_boxes']  if inference else ''
    inferred_weight = inference['inferred_weight'] if inference else ''
    inference_notes = inference['inference_notes'] if inference else ''
    packing_basis   = get_kit_pack_type(matched_id) if matched_id else ''

    if matched_id:
        matched_name   = kit_name_map.get(matched_id, str(matched_id))
        source_label   = comp_label(matched_id)
        score_pct      = f'{best_score:.0%}'
        output_records = select_output_records(matched_id, packing_basis, inferred_boxes, match_type)

        # Keep "Inferred Boxes" consistent with listed rows for partial Box outputs.
        if packing_basis == 'Box' and str(match_type).startswith('Partial'):
            try:
                predicted_boxes = int(inferred_boxes)
            except (TypeError, ValueError):
                predicted_boxes = None
            if predicted_boxes and output_records and len(output_records) < predicted_boxes:
                inference_notes = (
                    f'{inference_notes} | '
                    f'Output capped to {len(output_records)} listed source box rows '
                    f'(predicted {predicted_boxes}).'
                ).strip()
                inferred_boxes = len(output_records)

        for idx, ms in enumerate(output_records, start=1):
            row = _empty_row(kit_id, kit_name, my_label, match_type)
            row['Match Score']        = score_pct
            row['Matched Kit ID']     = matched_id
            row['Matched Kit Name']   = matched_name
            row['Source Components']  = source_label
            row['Date Created']       = ms['created']
            if packing_basis == 'Box' and str(match_type).startswith('Partial'):
                row['Package Number'] = idx
            else:
                row['Package Number'] = ms['pkg_number']
            row['Package Type']       = get_pkg_type(ms)
            row['Num Boxes in Pallet']= ms['num_boxes']
            row['Weight']             = ms['weight']
            row['Length']             = ms['length']
            row['Width']              = ms['width']
            row['Height']             = ms['height']
            if packing_basis == 'Box' and str(match_type).startswith('Partial'):
                row['Inferred Boxes'] = idx
                row['Total Inferred Boxes'] = inferred_boxes
            else:
                row['Inferred Boxes'] = inferred_boxes
                row['Total Inferred Boxes'] = inferred_boxes
            row['Inferred Weight']    = inferred_weight
            row['Packing Basis']      = packing_basis
            row['Inference Notes']    = inference_notes
            results.append(row)
    else:
        results.append(_empty_row(kit_id, kit_name, my_label, match_type))


# ── Summary ───────────────────────────────────────────────────────────────────
df = pd.DataFrame(results)
match_counts = df.drop_duplicates('Kit ID')['Match Type'].value_counts()
print("\nMatch summary:")
print(match_counts.to_string())


# ── Write Excel ───────────────────────────────────────────────────────────────
with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    summary_df = match_counts.reset_index()
    summary_df.columns = ['Match Type', 'Kit Count']
    summary_df.to_excel(writer, sheet_name='Summary', index=False)

    df[df['Match Type'] == 'Component Decomposition'].sort_values('Kit Name').to_excel(
        writer, sheet_name='Component Decomp', index=False)
    df[df['Match Type'] == 'Exact'].sort_values('Kit Name').to_excel(
        writer, sheet_name='Exact Matches', index=False)
    df[df['Match Type'].str.startswith('Partial', na=False)].sort_values('Kit Name').to_excel(
        writer, sheet_name='Partial Matches', index=False)
    df[df['Match Type'] == 'No match'].sort_values('Kit Name').to_excel(
        writer, sheet_name='No Match', index=False)
    df.sort_values(['Match Type', 'Kit Name']).to_excel(
        writer, sheet_name='All', index=False)

# ── Style ─────────────────────────────────────────────────────────────────────
wb = load_workbook(OUTPUT_FILE)
COLORS = {
    'Component Decomposition': 'BDD7EE',   # steel blue
    'Exact'                  : 'C6EFCE',   # green
    'Partial'                : 'FFEB9C',   # amber
    'No match'               : 'FFC7CE',   # red
    'header'                 : '1F4E79',
}
header_font = Font(color='FFFFFF', bold=True)

for ws in wb.worksheets:
    for cell in ws[1]:
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'

    col_names = [c.value for c in ws[1]]
    if 'Match Type' in col_names:
        mt_col = col_names.index('Match Type') + 1
        for row in ws.iter_rows(min_row=2):
            mt_val = row[mt_col - 1].value or ''
            if 'Decomp' in mt_val or 'Component' in mt_val:
                color = COLORS['Component Decomposition']
            elif 'Exact' in mt_val:
                color = COLORS['Exact']
            elif 'Partial' in mt_val:
                color = COLORS['Partial']
            elif 'No match' in mt_val:
                color = COLORS['No match']
            else:
                color = 'F2F2F2'
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            for cell in row:
                cell.fill = fill

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 70)

wb.save(OUTPUT_FILE)
print(f'\nDone! Output: {OUTPUT_FILE}')
